"""
Extrai dados dos cadernos de leite (2003-2026) para JSON normalizado.

3 formatos:
  - Format A (2003-2012, .xls): 1 linha/dia, 2 turnos (Manhã+Tarde), colunas offset por 0 ou 1
  - Format B (2013-2021, .xlsx): 1 linha/dia, 4 turnos (Noite+Madrugada+Manhã+Tarde)
  - Format C (2022-2026, .xlsx/.xlsm): 3 linhas/dia (Manhã+Tarde+Noite), múltiplos tanques

Uso: python3 extract-cadernos.py [--dry-run]
"""

import json
import os
import sys
import glob
from datetime import date
from typing import Optional

MONTH_NAMES = {
    'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'ABRIL': 4,
    'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8,
    'SETEMBRO': 9, 'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12,
    # Abbreviated forms used in 2022+ format
    'JAN': 1, 'FEV': 2, 'MAR': 3, 'ABR': 4,
    'MAI': 5, 'JUN': 6, 'JUL': 7, 'AGO': 8,
    'SET': 9, 'OUT': 10, 'NOV': 11, 'DEZ': 12,
}

SKIP_SHEETS = {
    'GRÁFICO', 'MODELO', 'LEITEVENCUR', 'LEITEGRANJA', 'PLAN1',
    'PLANILHA1', 'PLANILHA2', 'CADERNO', 'DADOS GRAF. DEZ2003',
}


def num(val) -> Optional[float]:
    if val is None or val == '' or val == 0:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def int_or_none(val) -> Optional[int]:
    n = num(val)
    return int(n) if n is not None else None


def detect_month_from_sheet(sheet_name: str, year: int) -> Optional[int]:
    """Detect month number from sheet name."""
    upper = sheet_name.upper().strip()

    # Try direct match
    for name, month in MONTH_NAMES.items():
        if name in upper:
            return month

    # Try abbreviated: "Jan 25", "Fev 24", etc.
    for name, month in MONTH_NAMES.items():
        if upper.startswith(name):
            return month

    return None


def detect_year_from_filename(filename: str) -> int:
    """Extract year from filename like caderno_leite_2013.xlsx"""
    base = os.path.basename(filename)
    for part in base.replace('.', '_').split('_'):
        if part.isdigit() and len(part) == 4:
            return int(part)
    raise ValueError(f"Cannot detect year from {filename}")


def detect_format(filename: str, year: int) -> str:
    """Detect which format parser to use."""
    if filename.endswith('.xls') and not filename.endswith('.xlsx'):
        return 'A'  # Old binary .xls
    if year <= 2021:
        return 'B'  # .xlsx with 1 row/day, 4 shifts
    return 'C'  # .xlsx/.xlsm with 3 rows/day, multiple tanks


# ─── Format A: 2003-2012 (.xls) ─────────────────────────────────

def parse_format_a(filename: str, year: int) -> list:
    import xlrd
    wb = xlrd.open_workbook(filename)
    records = []

    for sheet_name in wb.sheet_names():
        month = detect_month_from_sheet(sheet_name, year)
        if month is None:
            continue

        ws = wb.sheet_by_name(sheet_name)

        # Detect column offset: some years start at col 0, others at col 1
        # Find the header row with 'Data' to determine offset
        offset = 0
        data_start_row = 6  # default
        for r in range(10):
            for c in range(3):
                val = str(ws.cell_value(r, c)).strip()
                if val == 'Data':
                    offset = c
                    data_start_row = r + 1
                    break

        for r in range(data_start_row, ws.nrows):
            day_val = ws.cell_value(r, offset)
            if not day_val or not isinstance(day_val, (int, float)):
                continue
            day = int(day_val)
            if day < 1 or day > 31:
                continue

            try:
                d = date(year, month, day)
            except ValueError:
                continue

            # Column positions relative to offset:
            # Manhã: offset+1=Régua, offset+2=Litros
            # Tarde 1º tanque: offset+3=Régua, offset+4=Litros
            # Tarde 2º tanque: offset+5=Régua, offset+6=Litros
            # Total: offset+7
            # Vacas: offset+8
            # Média: offset+9
            # Caminhão: offset+10
            # Leiteira: offset+11

            morning_liters = num(ws.cell_value(r, offset + 2))
            afternoon_1 = num(ws.cell_value(r, offset + 4))
            afternoon_2 = num(ws.cell_value(r, offset + 6))
            afternoon_liters = None
            if afternoon_1 is not None or afternoon_2 is not None:
                afternoon_liters = (afternoon_1 or 0) + (afternoon_2 or 0)

            total = num(ws.cell_value(r, offset + 7))
            cow_count = int_or_none(ws.cell_value(r, offset + 8))
            avg = num(ws.cell_value(r, offset + 9))
            collection = num(ws.cell_value(r, offset + 10))
            nursery = num(ws.cell_value(r, offset + 11))

            if total is None or total <= 0:
                continue

            records.append({
                'date': d.isoformat(),
                'totalLiters': total,
                'cowCount': cow_count,
                'avgPerCow': round(avg, 2) if avg else None,
                'morningLiters': morning_liters,
                'afternoonLiters': afternoon_liters,
                'nightLiters': None,
                'dawnLiters': None,
                'collectionLiters': collection,
                'nurseryLiters': nursery,
                'discardLiters': None,
                'calfLiters': None,
                'source': os.path.basename(filename),
            })

    return records


# ─── Format B: 2013-2021 (.xlsx) ────────────────────────────────
# Layout varies per sheet! Possible layouts detected by header row:
#   B1: NOITE | MADRUGADA | MANHÃ | TARDE | TOTAL  (4 shifts, cols B-J)
#   B2: MANHÃ | TOTAL                               (1 shift, cols B-D)
#   B3: MANHÃ | TARDE | TOTAL                       (2 shifts, cols B-F)
#   B4: MANHÃ | TARDE | Noite | TOTAL               (3 shifts, cols B-H)

SHIFT_NAMES = {'MANHÃ', 'TARDE', 'NOITE', 'MADRUGADA'}


def detect_b_layout(ws, header_row: int) -> dict:
    """Detect column positions dynamically from the header row.

    Returns dict with keys: shifts (list of (name, litros_col)),
    total_col, vacas_col, media_col, caminhao_col, leiteira_col.
    """
    # Read header row values
    cols = {}
    for c in range(1, 20):
        val = ws.cell(header_row, c).value
        if val:
            cols[c] = str(val).strip().upper()

    # Find TOTAL column
    total_col = None
    for c, v in cols.items():
        if 'TOTAL' in v:
            total_col = c
            break

    if total_col is None:
        return None

    # Shifts are between col 2 and total_col
    # Each shift occupies 2 columns (Régua + Litros)
    shifts = []
    for c, v in sorted(cols.items()):
        if c >= total_col:
            break
        if c < 2:
            continue
        if v in SHIFT_NAMES:
            # Litros col is the next odd column (c+1 for the litros value)
            shifts.append((v, c + 1))

    # After TOTAL: Nº vacas, Média, Caminhão, Leiteira
    return {
        'shifts': shifts,
        'total_col': total_col,
        'vacas_col': total_col + 1,
        'media_col': total_col + 2,
        'caminhao_col': total_col + 3,
        'leiteira_col': total_col + 4,
    }


def parse_format_b(filename: str, year: int) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(filename, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        month = detect_month_from_sheet(sheet_name, year)
        if month is None:
            continue

        ws = wb[sheet_name]

        # Find header row with shift names (row that has MANHÃ/NOITE and TOTAL)
        header_row = None
        data_start_row = None
        for r in range(1, 10):
            val = ws.cell(r, 1).value
            if val and str(val).strip() == 'Data':
                data_start_row = r + 1
            # Check if this row has TOTAL
            for c in range(2, 16):
                cv = ws.cell(r, c).value
                if cv and 'TOTAL' in str(cv).upper():
                    header_row = r
                    break
            if header_row:
                break

        if header_row is None:
            continue
        if data_start_row is None:
            data_start_row = header_row + 2  # fallback

        layout = detect_b_layout(ws, header_row)
        if layout is None:
            continue

        for r in range(data_start_row, ws.max_row + 1):
            day_val = ws.cell(r, 1).value
            if not day_val or not isinstance(day_val, (int, float)):
                continue
            day = int(day_val)
            if day < 1 or day > 31:
                continue

            try:
                d = date(year, month, day)
            except ValueError:
                continue

            # Read shifts dynamically
            shift_values = {}
            for shift_name, litros_col in layout['shifts']:
                shift_values[shift_name] = num(ws.cell(r, litros_col).value)

            total = num(ws.cell(r, layout['total_col']).value)
            cow_count = int_or_none(ws.cell(r, layout['vacas_col']).value)
            avg = num(ws.cell(r, layout['media_col']).value)
            collection = num(ws.cell(r, layout['caminhao_col']).value)
            nursery = num(ws.cell(r, layout['leiteira_col']).value)

            if total is None or total <= 0:
                continue

            records.append({
                'date': d.isoformat(),
                'totalLiters': total,
                'cowCount': cow_count,
                'avgPerCow': round(avg, 2) if avg else None,
                'morningLiters': shift_values.get('MANHÃ'),
                'afternoonLiters': shift_values.get('TARDE'),
                'nightLiters': shift_values.get('NOITE'),
                'dawnLiters': shift_values.get('MADRUGADA'),
                'collectionLiters': collection,
                'nurseryLiters': nursery,
                'discardLiters': None,
                'calfLiters': None,
                'source': os.path.basename(filename),
            })

    return records


# ─── Format C: 2022-2026 (.xlsx/.xlsm) ──────────────────────────

def parse_format_c(filename: str, year: int) -> list:
    import openpyxl
    wb = openpyxl.load_workbook(filename, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper().strip()
        if upper in SKIP_SHEETS:
            continue

        month = detect_month_from_sheet(sheet_name, year)
        if month is None:
            continue

        ws = wb[sheet_name]

        # Format C: 3 rows per day (Manhã, Tarde, Noite)
        # Find data start row (after headers)
        data_start_row = 4  # typically row 4
        for r in range(1, 8):
            val = ws.cell(r, 1).value
            if val == 'Dia':
                data_start_row = r + 2  # skip Dia + sub-header rows
                break

        # Group rows by day
        current_day = None
        day_data = {}

        for r in range(data_start_row, ws.max_row + 1):
            day_cell = ws.cell(r, 1).value
            shift_cell = ws.cell(r, 2).value

            if day_cell is not None and isinstance(day_cell, (int, float)):
                current_day = int(day_cell)
                if current_day < 1 or current_day > 31:
                    current_day = None
                    continue
                if current_day not in day_data:
                    day_data[current_day] = {
                        'morning': 0, 'afternoon': 0, 'night': 0,
                        'collection': 0, 'nursery': 0, 'discard': 0, 'calf': 0,
                        'total_day': 0, 'cow_count': None, 'avg': None,
                    }

            if current_day is None or current_day not in day_data:
                continue

            shift = str(shift_cell or '').strip().lower()
            if shift not in ('manhã', 'tarde', 'noite'):
                continue

            dd = day_data[current_day]

            # P = Leite Total Ordenha (per shift)
            shift_liters = num(ws.cell(r, 16).value) or 0

            if shift == 'manhã':
                dd['morning'] += shift_liters
            elif shift == 'tarde':
                dd['afternoon'] += shift_liters
            elif shift == 'noite':
                dd['night'] += shift_liters

            # L = Total Caminhão (collection pickup)
            coll = num(ws.cell(r, 12).value)
            if coll:
                dd['collection'] += coll

            # M = Leiteira, N = Bezerras, O = Descarte
            nurs = num(ws.cell(r, 13).value)
            if nurs:
                dd['nursery'] += nurs

            calf = num(ws.cell(r, 14).value)
            if calf:
                dd['calf'] += calf

            disc = num(ws.cell(r, 15).value)
            if disc:
                dd['discard'] += disc

            # Q = Total do dia (only on first row of day)
            total_day = num(ws.cell(r, 17).value)
            if total_day and total_day > 0:
                dd['total_day'] = total_day

            # R = Vacas
            cows = int_or_none(ws.cell(r, 18).value)
            if cows and cows > 0:
                dd['cow_count'] = cows

            # U = Média (only on some sheets)
            avg = num(ws.cell(r, 21).value)
            if avg and avg > 0:
                dd['avg'] = round(avg, 2)

        # Convert day_data to records
        for day, dd in sorted(day_data.items()):
            try:
                d = date(year, month, day)
            except ValueError:
                continue

            total = dd['total_day']
            if total <= 0:
                # Fall back to sum of shifts
                total = dd['morning'] + dd['afternoon'] + dd['night']
            if total <= 0:
                continue

            records.append({
                'date': d.isoformat(),
                'totalLiters': total,
                'cowCount': dd['cow_count'],
                'avgPerCow': dd['avg'] or (round(total / dd['cow_count'], 2) if dd['cow_count'] else None),
                'morningLiters': dd['morning'] or None,
                'afternoonLiters': dd['afternoon'] or None,
                'nightLiters': dd['night'] or None,
                'dawnLiters': None,
                'collectionLiters': dd['collection'] or None,
                'nurseryLiters': dd['nursery'] or None,
                'discardLiters': dd['discard'] or None,
                'calfLiters': dd['calf'] or None,
                'source': os.path.basename(filename),
            })

    return records


# ─── Main ────────────────────────────────────────────────────────

def main():
    dry_run = '--dry-run' in sys.argv

    # Find all caderno files in project root
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..', '..', '..'))
    pattern = os.path.join(project_root, 'caderno_leite_*')
    files = sorted(glob.glob(pattern))

    if not files:
        print(f'No caderno_leite_* files found in {project_root}')
        sys.exit(1)

    all_records = []
    seen_dates = set()

    for filepath in files:
        filename = os.path.basename(filepath)
        # Skip _old variants if we have the main file
        if '_old' in filename:
            print(f'  ⊘ Skipping {filename} (old variant)')
            continue

        year = detect_year_from_filename(filepath)
        fmt = detect_format(filepath, year)

        print(f'  → {filename} (year={year}, format={fmt})')

        try:
            if fmt == 'A':
                records = parse_format_a(filepath, year)
            elif fmt == 'B':
                records = parse_format_b(filepath, year)
            else:
                records = parse_format_c(filepath, year)
        except Exception as e:
            print(f'    ✗ Error: {e}')
            continue

        # Deduplicate: keep first occurrence per date
        new_count = 0
        for rec in records:
            if rec['date'] not in seen_dates:
                seen_dates.add(rec['date'])
                all_records.append(rec)
                new_count += 1

        print(f'    ✓ {len(records)} rows extracted, {new_count} new (after dedup)')

    # Sort by date
    all_records.sort(key=lambda r: r['date'])

    print(f'\n  Total: {len(all_records)} records ({all_records[0]["date"]} → {all_records[-1]["date"]})')

    if dry_run:
        # Print sample
        print('\n  Sample (first 5):')
        for rec in all_records[:5]:
            print(f'    {rec["date"]}: {rec["totalLiters"]}L, {rec["cowCount"]} vacas, col={rec["collectionLiters"]}')
        print('\n  Sample (last 5):')
        for rec in all_records[-5:]:
            print(f'    {rec["date"]}: {rec["totalLiters"]}L, {rec["cowCount"]} vacas, col={rec["collectionLiters"]}')

        # Stats per year
        print('\n  Per year:')
        years = {}
        for rec in all_records:
            y = rec['date'][:4]
            if y not in years:
                years[y] = {'count': 0, 'total': 0}
            years[y]['count'] += 1
            years[y]['total'] += rec['totalLiters']
        for y in sorted(years.keys()):
            print(f'    {y}: {years[y]["count"]} days, {years[y]["total"]:,.0f} L total')
    else:
        # Write JSON
        output_path = os.path.join(os.path.dirname(__file__), 'cadernos-leite.json')
        with open(output_path, 'w') as f:
            json.dump(all_records, f, indent=2, ensure_ascii=False)
        print(f'\n  → JSON written to {output_path}')


if __name__ == '__main__':
    main()
